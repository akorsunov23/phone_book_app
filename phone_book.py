import os.path

import openpyxl


class PhoneBook:
    """Класс телефонного справочника."""

    def __init__(self):
        """
        Инициализация справочника.
        Если его нет, тогда создаём.
        """
        self._filename: str = "phonebook.xlsx"
        if os.path.exists(self._filename):
            self._book = openpyxl.load_workbook(
                filename=self._filename
            )
            self._row = self._book.active
        else:
            self._book = openpyxl.Workbook()
            self._row = self._book.active
            headings: list = [
                "ФИО",
                "Организация",
                "Телефон рабочий",
                "Телефон личный"
            ]
            row_num: int = 1
            col_num: int = 1
            for heading in headings:
                self._row.cell(
                    row=row_num,
                    column=col_num,
                    value=heading
                )
                col_num += 1
            self._book.save(filename=self._filename)


class ActionBook(PhoneBook):
    """Действия над справочником."""

    def read_book(self):
        """Чтение справочника."""
        if self._row.max_row > 1:
            for index, row in enumerate(self._row):
                if index != 0:
                    print(
                        f"\t{index}. {' '.join([v.value for v in row])}"
                    )
            return self._row.max_row
        print("Справочник пуст.")

    def add_val_book(self, values: list):
        """Добавление записи в справочник."""

        try:
            data: list = [(val.title() for val in values)]
            row_num = self._row.max_row + 1
            for row in data:
                col_num: int = 1
                for cell in row:
                    self._row.cell(
                        row=row_num,
                        column=col_num,
                        value=cell
                    )
                    col_num += 1
                row_num += 1
            self._book.save(filename=self._filename)
            print(f"\n{values[0]} успешно добавлен.")
        except Exception as ex:
            print(f"Произошла ошибка при добавлении {ex}")

    def update_val_book(self, values: dict):
        """Обновление записи в справочнике."""

        try:
            self._row.cell(
                row=values["row"],
                column=values["field"]
            ).value = values[
                "value"
            ]
            self._book.save(filename=self._filename)
            print("Запись успешно изменена.")
        except Exception as ex:
            print(f"Произошла ошибка при редактировании {ex}")

    def search_val_book(self, values: list):
        """Поиск записи в справочнике."""

        response: list = []

        for i_val, val in values:
            for i_row, row in enumerate(self._row):
                if i_row != 0:
                    record: list = [v.value for v in row]
                    if record[i_val] == val and record not in response:
                        response.append(record)

        if len(response) > 0:
            print("Найденные записи:\n")
            for value in response:
                print(" ".join(value))
        else:
            print("\nНичего не найдено!")


book = ActionBook()
